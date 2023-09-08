import pandas as pd
import streamlit as st
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, NamedStyle, Font
from datetime import datetime
from openpyxl.styles import Alignment


# Convert DataFrame to formatted Excel bytes
def convert_df_to_excel(df, myheader, subheader):
    wb = Workbook()
    ws = wb.active

    # Styles
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    header_font = Font(name='Calibri', size=14, bold=True)
    sub_header_font = Font(name='Calibri', size=12, bold=True)

    gray_style = NamedStyle(name="gray", fill=gray_fill, border=thin_border)
    white_style = NamedStyle(name="white", fill=white_fill, border=thin_border)
    header_style = NamedStyle(name="header", font=header_font, alignment=Alignment(horizontal='center'))
    sub_header_style = NamedStyle(name="sub_header", font=sub_header_font, alignment=Alignment(horizontal='center'))

    wb.add_named_style(gray_style)
    wb.add_named_style(white_style)
    wb.add_named_style(header_style)
    wb.add_named_style(sub_header_style)

    # Add a header
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df.shape[1] + 1)  # Assuming the index is included
    ws['A1'].value = myheader
    ws['A1'].style = "header"

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=df.shape[1] + 1)
    ws['A2'].value = subheader
    ws['A2'].style = "sub_header"

    # Write the DataFrame to Excel, including the index
    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 3):  # Start from row 3 due to the header
        row_style = gray_style if (r_idx - 2) % 2 == 0 else white_style
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.style = row_style.name

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return stream


def df_to_excel(df, myheader='Relatório de Dados', subheader=''):
    if st.button('gerar relatório em arquivo de excel'):
        excel_bytes = convert_df_to_excel(df, myheader, subheader)
        st.download_button(
            label="baixar relatório em arquivo de excel",
            data=excel_bytes,
            file_name='data.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
